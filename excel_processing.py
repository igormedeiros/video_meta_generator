import os

from openpyxl import Workbook, load_workbook

from file_processing import save_video_copy
from logger import log


def save_to_excel(output_file, video_file, title, description, hashtags):
    log.debug(f'save_to_excel({output_file}, '
              f'{video_file}, {title}, {description}, {hashtags})')
    """Saves video metadata to an Excel file and includes the copied video file."""

    # Create or load the Excel workbook
    if not os.path.exists(output_file):
        wb = Workbook()
        ws = wb.active
        ws.append(["File Path", "Title", "Description", "Hashtags"])
    else:
        wb = load_workbook(output_file)
        ws = wb.active

    # Save a copy of the video with the determined title in the 'videos_processed' folder
    processed_video_file = save_video_copy(video_file, title)

    ws.append([os.path.abspath(processed_video_file), title, description, hashtags])

    # Save the workbook to the output file
    wb.save(output_file)
