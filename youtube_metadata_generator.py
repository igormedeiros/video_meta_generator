
import openai
import os
import sys
from openpyxl import Workbook, load_workbook
import json

# Load configurations from the provided config file
with open("config.json", "r") as config_file:
    config = json.load(config_file)

openai.api_key = config["openai_api_key"]


def generate_youtube_metadata(transcript):
    """Generate YouTube title, description, and hashtags using the OpenAI API."""
    prompt_text = f'{config["prompt"]}{transcript}'
    response = openai.Completion.create(
        engine=config["engine"],
        prompt=prompt_text,
        max_tokens=config["max_tokens"]
    )
    output = response.choices[0].text.strip().split('\n')
    title = next(line.replace("Title: ", "").strip()
                 for line in output if "Title:" in line)
    description = next(line.replace("Description: ", "").strip()
                       for line in output if "Description:" in line)
    hashtags = next(line.replace("Hashtags: ", "").strip()
                    for line in output if "Hashtags:" in line)
    return title, description, hashtags


def save_to_spreadsheet(file_path, title, description, hashtags, output_file):
    """Save the generated metadata to the provided XLSX spreadsheet."""
    if not os.path.exists(output_file):
        wb = Workbook()
        ws = wb.active
        ws.append(["File Path", "Title", "Description", "Hashtags"])
    else:
        wb = load_workbook(output_file)
        ws = wb.active
    ws.append([file_path, title, description, hashtags])
    wb.save(output_file)


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print(
            "Usage: python script_name.py [path_to_transcript.txt] [output.xlsx]")
        sys.exit(1)
    file_path = sys.argv[1]
    output_file = sys.argv[2]
    with open(file_path, 'r', encoding='utf-8') as file:
        transcript = file.read()
    title, description, hashtags = generate_youtube_metadata(transcript)
    print(f"Generated Title: {title}")
    print(f"Generated Description: {description}")
    print(f"Generated Hashtags: {hashtags}")
    save_to_spreadsheet(file_path, title, description, hashtags, output_file)
    print(f"Data saved to {output_file}")
